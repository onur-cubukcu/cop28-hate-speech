"""
COP28 Hate Speech Detection Pipeline
=====================================
Reads scraped tweets (from Apify), classifies each one, writes results
to Excel, and produces charts.

Classification approach (faithful to original Toraman22 methodology):
  Step 1 — cardiffnlp/twitter-roberta-base-offensive
           Trained on Twitter data. Detects offensive language broadly
           (profanity, insults, angry language).
           Labels: offensive / not-offensive

  Step 2 — Violence keyword check on offensive tweets
           If an offensive tweet contains threat/violence keywords
           it is escalated to "Hate", otherwise stays "Offensive".

  Final labels:
      0 → Neutral    (not offensive)
      1 → Offensive  (offensive but no violence keywords)
      2 → Hate       (offensive + violence/threat keywords)

This mirrors the original study's description:
  "the AI module only detects words such as swear words ('idiot') or
   words associated with violence ('kill', 'shoot')"

Usage
-----
    python pipeline.py --input tweets.json --output results.xlsx
"""

import argparse
import json
import re
from pathlib import Path

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

MODEL_NAME = "cardiffnlp/twitter-roberta-base-offensive"

# Keywords that escalate offensive → hate (threats/violence)
VIOLENCE_KEYWORDS = re.compile(
    r"\b(kill|shoot|murder|bomb|attack|lynch|hang|stab|gun|shoot|die|death threat|"
    r"destroy|eliminate|exterminate|execute|slaughter|massacre|hurt|harm|"
    r"beat|punch|rape|assault)\b",
    re.IGNORECASE
)

COLORS = {
    "Neutral":   "#4CAF50",
    "Offensive": "#FF9800",
    "Hate":      "#F44336",
}

FILL_COLORS = {
    "Neutral":   "C8E6C9",
    "Offensive": "FFE0B2",
    "Hate":      "FFCDD2",
}


def load_tweets(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    rows = []
    for item in data:
        text = item.get("full_text") or item.get("text") or ""
        if item.get("handle"):
            author = item["handle"]
        elif isinstance(item.get("user"), dict):
            author = item["user"].get("screen_name", "")
        else:
            author = item.get("author_id", "")

        rows.append({
            "id":         item.get("id_str") or item.get("id", ""),
            "created_at": item.get("created_at", ""),
            "author":     author,
            "text":       text.strip(),
            "retweets":   item.get("retweet_count", 0),
            "likes":      item.get("favorite_count", 0),
        })

    df = pd.DataFrame(rows)
    df = df[df["text"].str.len() > 0].reset_index(drop=True)
    print(f"[load]  Loaded {len(df)} tweets from '{path}'")
    return df


def load_classifier():
    from transformers import pipeline as hf_pipeline
    print(f"[model] Loading '{MODEL_NAME}'")
    print("[model] First run downloads the model - please wait...")
    clf = hf_pipeline(
        "text-classification",
        model=MODEL_NAME,
        truncation=True,
        max_length=512,
    )
    print("[model] Model ready.")
    return clf


def classify_all(df, clf):
    texts = df["text"].tolist()
    total = len(texts)
    ids, names, scores = [], [], []

    print(f"[classify] Processing {total} tweets...")
    for i, text in enumerate(texts, start=1):
        try:
            result    = clf(text[:512])[0]
            raw_label = result["label"].lower()   # "offensive" or "non-offensive"
            score     = round(result["score"], 4)

            if "non" in raw_label or raw_label == "not-offensive":
                # Not offensive at all → Neutral
                lid, lname = 0, "Neutral"
            else:
                # Offensive — check for violence keywords to escalate to Hate
                if VIOLENCE_KEYWORDS.search(text):
                    lid, lname = 2, "Hate"
                else:
                    lid, lname = 1, "Offensive"

        except Exception as e:
            print(f"  [warn] Error on tweet {i}: {e}")
            lid, lname, score = 0, "Neutral", 0.0

        ids.append(lid)
        names.append(lname)
        scores.append(score)

        if i % 100 == 0 or i == total:
            print(f"  [{i}/{total}] done...")

    df["label_id"]   = ids
    df["label_name"] = names
    df["confidence"] = scores
    return df


def save_excel(df, output_path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Results", index=False)
        counts = df["label_name"].value_counts()
        summary = pd.DataFrame({
            "Category":   counts.index,
            "Count":      counts.values,
            "Percentage": (counts.values / len(df) * 100).round(2),
        })
        summary.to_excel(writer, sheet_name="Summary", index=False)

    wb = load_workbook(output_path)
    ws_res = wb["Results"]
    col_idx = df.columns.tolist().index("label_name") + 1

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="37474F")
    for cell in ws_res[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in ws_res.iter_rows(min_row=2):
        label_val = row[col_idx - 1].value or "Neutral"
        fill = PatternFill("solid", fgColor=FILL_COLORS.get(label_val, "FFFFFF"))
        for cell in row:
            cell.fill = fill

    for ws in [ws_res, wb["Summary"]]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 80)

    wb.save(output_path)
    print(f"[excel] Saved -> '{output_path}'")


def make_charts(df, output_dir, tag="#COP28"):
    out = Path(output_dir)
    out.mkdir(exist_ok=True)

    counts = df["label_name"].value_counts().reindex(["Neutral", "Offensive", "Hate"], fill_value=0)
    labels = counts.index.tolist()
    values = counts.values
    pct    = values / values.sum() * 100
    colors = [COLORS[l] for l in labels]

    # Pie chart
    fig, ax = plt.subplots(figsize=(7, 7))
    wedges, _, autotexts = ax.pie(
        values, labels=None, autopct="%1.2f%%", colors=colors,
        startangle=90, wedgeprops=dict(edgecolor="white", linewidth=2),
        pctdistance=0.75,
    )
    for at in autotexts:
        at.set_fontsize(13)
        at.set_fontweight("bold")
    legend_labels = [f"{l}  ({v:,}  |  {p:.2f}%)" for l, v, p in zip(labels, values, pct)]
    ax.legend(wedges, legend_labels, loc="lower center",
              bbox_to_anchor=(0.5, -0.08), ncol=1, frameon=False, fontsize=11)
    ax.set_title(f"Hate Speech on X - search criteria: {tag}", fontsize=14, fontweight="bold", pad=20)
    pie_path = out / "pie_chart.png"
    fig.savefig(pie_path, bbox_inches="tight", dpi=150)
    plt.close(fig)
    print(f"[chart] Pie chart -> '{pie_path}'")

    # Bar chart
    fig, ax = plt.subplots(figsize=(7, 5))
    bars = ax.bar(labels, pct, color=colors, edgecolor="white", linewidth=1.5, width=0.5)
    for bar, p in zip(bars, pct):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                f"{p:.2f}%", ha="center", va="bottom", fontweight="bold", fontsize=12)
    ax.set_ylabel("Percentage (%)", fontsize=12)
    ax.set_title(f"Classification Breakdown - {tag}", fontsize=13, fontweight="bold")
    ax.set_ylim(0, max(pct) * 1.15)
    ax.spines[["top", "right"]].set_visible(False)
    bar_path = out / "bar_chart.png"
    fig.savefig(bar_path, bbox_inches="tight", dpi=150)
    plt.close(fig)
    print(f"[chart] Bar chart -> '{bar_path}'")


def main():
    parser = argparse.ArgumentParser(description="COP28 Hate-Speech Detection Pipeline")
    parser.add_argument("--input",  required=True,          help="Path to Apify JSON export")
    parser.add_argument("--output", default="results.xlsx", help="Output Excel file")
    parser.add_argument("--charts", default="charts",       help="Folder for chart images")
    parser.add_argument("--tag",    default="#COP28",       help="Label shown in chart titles")
    args = parser.parse_args()

    df  = load_tweets(args.input)
    clf = load_classifier()
    df  = classify_all(df, clf)
    save_excel(df, args.output)
    make_charts(df, args.charts, tag=args.tag)

    print("\n=== Summary ===")
    counts = df["label_name"].value_counts()
    for cat, n in counts.items():
        print(f"  {cat:<12} {n:>5}  ({n/len(df)*100:.2f} %)")
    print(f"  {'TOTAL':<12} {len(df):>5}")


if __name__ == "__main__":
    main()
